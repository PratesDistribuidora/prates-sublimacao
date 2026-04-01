"""
importador.py — Prates Sublimação
Importa dados da planilha Excel original para o banco SQLite.
"""

from openpyxl import load_workbook
from database import (
    get_conn, upsert_sku, add_historico,
    FORNECEDORES_PADRAO
)


def importar_xlsx(caminho):
    """
    Importa dados do xlsx original para o banco.
    Retorna dict com resumo do que foi importado.
    """
    try:
        wb = load_workbook(caminho, read_only=False, data_only=True)
    except Exception as e:
        return {'erro': str(e)}

    resumo = {'skus': 0, 'fornecedores': 0, 'parametros': 0, 'erros': []}

    # ── SKUs (SUPER_REVENDA) ─────────────────────────────
    if 'SUPER_REVENDA' in wb.sheetnames:
        ws = wb['SUPER_REVENDA']
        for row in ws.iter_rows(min_row=2, values_only=True):
            modelo, tecido, cor, tamanho, _, peso_g = (
                row[0], row[1], row[2], row[3], row[4], row[5])
            if all([modelo, tecido, cor, tamanho]) and peso_g:
                try:
                    upsert_sku(modelo, tecido, cor, tamanho, float(peso_g))
                    resumo['skus'] += 1
                except Exception as e:
                    resumo['erros'].append(f"SKU {modelo}/{tecido}/{cor}/{tamanho}: {e}")

    # ── Fornecedores ─────────────────────────────────────
    if 'FORNECEDORES' in wb.sheetnames:
        ws = wb['FORNECEDORES']
        conn = get_conn()
        for row in ws.iter_rows(min_row=5, values_only=True):
            tecido, cor, chave = row[0], row[1], row[2]
            f1_nome, f1_preco = row[3], row[4]
            f2_nome, f2_preco = row[5], row[6]
            f3_nome, f3_preco = row[7], row[8]
            forn_ativo = row[10]
            if not tecido or not cor:
                continue
            try:
                conn.execute("""INSERT INTO fornecedores
                    (tecido,cor,chave,f1_nome,f1_preco,f2_nome,f2_preco,
                     f3_nome,f3_preco,fornecedor_ativo)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(chave) DO UPDATE SET
                        f1_nome=excluded.f1_nome, f1_preco=excluded.f1_preco,
                        f2_nome=excluded.f2_nome, f2_preco=excluded.f2_preco,
                        f3_nome=excluded.f3_nome, f3_preco=excluded.f3_preco,
                        fornecedor_ativo=excluded.fornecedor_ativo""",
                    (tecido, cor, f"{tecido}|{cor}",
                     f1_nome or 'Importline', float(f1_preco or 0),
                     f2_nome or 'Fornecedor 2', float(f2_preco or 0),
                     f3_nome or 'Fornecedor 3', float(f3_preco or 0),
                     forn_ativo or 'Mais Barato'))
                resumo['fornecedores'] += 1
            except Exception as e:
                resumo['erros'].append(f"Fornecedor {tecido}/{cor}: {e}")
        conn.commit()
        conn.close()

    # ── Faccionistas ─────────────────────────────────────
    if 'FACCIONISTAS' in wb.sheetnames:
        ws = wb['FACCIONISTAS']
        conn = get_conn()
        for row in ws.iter_rows(min_row=5, max_row=8, values_only=True):
            modelo = row[0]
            f1_nome, f1_preco = row[1], row[2]
            f2_nome, f2_preco = row[3], row[4]
            f3_nome, f3_preco = row[5], row[6]
            fac_ativa = row[8]
            if not modelo:
                continue
            try:
                conn.execute("""INSERT INTO faccionistas
                    (modelo,f1_nome,f1_preco,f2_nome,f2_preco,f3_nome,f3_preco,faccionista_ativa)
                    VALUES (?,?,?,?,?,?,?,?)
                    ON CONFLICT(modelo) DO UPDATE SET
                        f1_nome=excluded.f1_nome, f1_preco=excluded.f1_preco,
                        f2_nome=excluded.f2_nome, f2_preco=excluded.f2_preco,
                        f3_nome=excluded.f3_nome, f3_preco=excluded.f3_preco,
                        faccionista_ativa=excluded.faccionista_ativa""",
                    (modelo,
                     f1_nome or 'Faccionista 1', float(f1_preco or 0),
                     f2_nome or 'Faccionista 2', float(f2_preco or 0),
                     f3_nome or 'Faccionista 3', float(f3_preco or 0),
                     fac_ativa or 'Mais Barata'))
            except Exception as e:
                resumo['erros'].append(f"Faccionista {modelo}: {e}")
        conn.commit()
        conn.close()

    # ── Parâmetros (LISTAS) ───────────────────────────────
    if 'LISTAS' in wb.sheetnames:
        ws = wb['LISTAS']
        mapa = {
            'Costura Branco – fallback (R$)': 'costura_branco',
            'Costura Outras – fallback (R$)': 'costura_outras',
            'Frete (%)': 'frete_pct',
            'Outros Custos (%)': 'outros_pct',
            'Embalagem por peça (R$)': 'embalagem',
            'Margem Super Revenda (%)': 'margem_sr',
            'Margem Atacado (%)': 'margem_atacado',
            'Margem Varejo (%)': 'margem_varejo',
        }
        conn = get_conn()
        for row in ws.iter_rows(min_row=2, max_row=15, values_only=True):
            nome_param = row[5]
            valor_param = row[6]
            if nome_param in mapa and valor_param is not None:
                try:
                    conn.execute('INSERT OR REPLACE INTO parametros VALUES (?,?)',
                                 (mapa[nome_param], float(valor_param)))
                    resumo['parametros'] += 1
                except Exception as e:
                    resumo['erros'].append(f"Parâmetro {nome_param}: {e}")
        conn.commit()
        conn.close()

    # ── Histórico (HISTORICO) ─────────────────────────────
    if 'HISTORICO' in wb.sheetnames:
        ws = wb['HISTORICO']
        conn = get_conn()
        for row in ws.iter_rows(min_row=4, values_only=True):
            data, tipo, tec_mod, p_ant, p_nov, _, forn, motivo = (
                row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7])
            if not data or not tipo:
                continue
            try:
                variacao = round((float(p_nov) - float(p_ant)) / float(p_ant), 4) if p_ant else 0
                conn.execute("""INSERT INTO historico
                    (data,tipo,tecido_modelo,preco_anterior,preco_novo,
                     variacao_pct,fornecedor_faccionista,motivo)
                    VALUES (?,?,?,?,?,?,?,?)""",
                    (str(data)[:10], tipo, tec_mod, float(p_ant or 0),
                     float(p_nov or 0), variacao, forn or '', motivo or ''))
            except Exception as e:
                resumo['erros'].append(f"Histórico linha: {e}")
        conn.commit()
        conn.close()

    return resumo
